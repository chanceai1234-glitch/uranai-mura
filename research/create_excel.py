# -*- coding: utf-8 -*-
"""01_uranai_list.md + 02_uranai_logic.md + 04_output_type.md を結合し uranai_master.xlsx を生成する。

依存: pip install openpyxl

04 は node research/build_output_type.mjs で再生成可能。
同内容を Node で生成する場合: npm install --prefix research exceljs; node research/create_excel.mjs
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
FILE_01 = ROOT / "01_uranai_list.md"
FILE_02 = ROOT / "02_uranai_logic.md"
FILE_04 = ROOT / "04_output_type.md"
OUT_XLSX = ROOT / "uranai_master.xlsx"

HEADERS = [
    "No.",
    "占い名",
    "発祥地・文化圏",
    "カテゴリ",
    "必要な入力情報",
    "判定ロジック概要",
    "出力内容",
    "更新頻度",
    "実装難易度",
    "出力方式",
    "出力方式詳細",
    "実装メモ",
    "概要メモ",
]

HEADER_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FONT = Font(name="Meiryo UI", size=11)
FONT_BOLD = Font(name="Meiryo UI", size=11, bold=True)
ALIGN_WRAP_TOP = Alignment(vertical="top", wrap_text=True)
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_table_rows(body: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for line in body.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) < 2:
            continue
        key, val = cells[0], cells[1]
        if key in ("項目", "------", ""):
            continue
        out[key] = val
    return out


def parse_01(text: str) -> list[dict[str, str]]:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    chunks = text.split("\n##### ")[1:]
    entries = []
    for ch in chunks:
        first_nl = ch.find("\n")
        if first_nl == -1:
            body = ""
        else:
            body = ch[first_nl + 1 :]
        d = parse_table_rows(body)
        entries.append(d)
    return entries


def parse_02(text: str) -> list[dict[str, str]]:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = text.split("\n### ")[1:]
    entries = []
    for p in parts:
        first_nl = p.find("\n")
        body = p[first_nl + 1 :] if first_nl != -1 else ""
        d = parse_table_rows(body)
        entries.append(d)
    return entries


def parse_04(text: str) -> dict[str, tuple[str, str]]:
    """| 占い名 | 出力方式 | 詳細 | → { 占い名: (方式, 詳細) }"""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    out: dict[str, tuple[str, str]] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        if "占い名" in line and "出力方式" in line:
            continue
        if re.match(r"^\|\s*[-:|]+\s*\|", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) < 2:
            continue
        name, typ = cells[0], cells[1]
        det = cells[2] if len(cells) > 2 else ""
        if not name:
            continue
        out[name] = (typ, det)
    return out


def col_width_units(text: str, max_w: int = 50) -> float:
    """日本語混在でも概算。Excel列幅の上限 max_w。"""
    if not text:
        return 8.0
    w = 0.0
    for ch in str(text):
        if ord(ch) < 128:
            w += 1.0
        else:
            w += 2.0
    return min(max(w * 0.9 + 2, 8), max_w)


def main() -> None:
    t1 = FILE_01.read_text(encoding="utf-8")
    t2 = FILE_02.read_text(encoding="utf-8")
    t4 = FILE_04.read_text(encoding="utf-8")

    list_01 = parse_01(t1)
    list_02 = parse_02(t2)
    out_type = parse_04(t4)

    if len(list_01) != len(list_02):
        raise SystemExit(
            f"件数不一致: 01={len(list_01)} 02={len(list_02)}。同じ並びで揃えてください。"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "占い一覧"

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = FONT_BOLD
        c.fill = HEADER_FILL
        c.alignment = ALIGN_HEADER

    for i, (d1, d2) in enumerate(zip(list_01, list_02), start=1):
        r = i + 1
        name = d2.get("占い名") or d1.get("占い名") or ""
        ot = out_type.get(name, ("", ""))
        if name and name not in out_type:
            print(f"警告: 04_output_type.md に占い名がありません: {name!r}")
        row_vals = [
            i,
            name,
            d1.get("発祥地・文化圏", ""),
            d1.get("出力カテゴリ", ""),
            d2.get("必要な入力情報", ""),
            d2.get("判定ロジック概要", ""),
            d2.get("出力内容", ""),
            d2.get("更新頻度", ""),
            d2.get("実装難易度", ""),
            ot[0],
            ot[1] if ot[0] == "C" else "",
            d2.get("実装メモ", ""),
            d1.get("概要メモ", ""),
        ]

        is_alt = (r % 2 == 0)
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = FONT
            c.alignment = ALIGN_WRAP_TOP
            if is_alt:
                c.fill = ALT_ROW_FILL

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        maxlen = 8.0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            maxlen = max(maxlen, col_width_units(v if v is not None else ""))
        ws.column_dimensions[letter].width = maxlen

    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({ws.max_row - 1} data rows)")


if __name__ == "__main__":
    main()
